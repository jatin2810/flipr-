from django.shortcuts import render, redirect
from itertools import islice
import os
import openpyxl
from .models import JobPost
from django.http import HttpResponse
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

# Create your views here.
def home(request):
    data = JobPost.objects.all()
    # user_list = User.objects.all()
    page = request.GET.get('page')

    paginator = Paginator(data, 10)
    try:
        context = paginator.page(page)
    except PageNotAnInteger:
        context = paginator.page(1)
    except EmptyPage:
        context = paginator.page(paginator.num_pages)

    # context={'data':data}
    return render(request, 'jobs/index.html',{'context':context})

def profile(request):
    return render(request, 'jobs/about.html')

def about(request):
    return render(request, 'jobs/profile.html')

def detail(request,pk):
    job = JobPost.objects.filter(id=pk).first()
    print(job)
    return render(request,'jobs/detail.html',{'job':job})

# def import_view(request):
#     cwd = os.getcwd()
#     file_to_read = cwd+'\\file\\sample.xlsx'
#     f = openpyxl.load_workbook(file_to_read)
#     sheet_obj = f.active
#     max_row = sheet_obj.max_row
#     max_col =sheet_obj.max_column
    
#     fList= []
#     for i in range(2, max_row):
#         templist = []
#         for j in range(1,max_col):
#             templist.append(sheet_obj.cell(row=i,column=j).value)
#         fList.append(templist)
   
#     batch_size = 1000
#     objs = (JobPost(company=data[0],education=data[1],experience=data[2],industry=data[3],jobdescription=data[4],jobid=data[5],joblocation_address=data[6],jobtitle=data[7],numberofpositions=data[8],payrate=data[9],postdate=data[10],site_name=data[11],skills=data[12],uniq_id=data[13]) for data in fList)
#     while True:
#         batch = list(islice(objs, batch_size))
#         if not batch:
#             break
#         JobPost.objects.bulk_create(batch, batch_size)
    
#     return HttpResponse('Success')
# import_view()
#         
